"""Unit tests for the pure core: scoring, extraction, link tokens, drafting, export.

Run either way:
    cd "Candidate screening" && python -m pytest tests -q
    cd "Candidate screening" && python tests/test_core.py
"""
from __future__ import annotations

import io
import os
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from backend import ai_agent, config, excel_export, extractors, interview_link, scoring  # noqa: E402


# ------------------------------------------------------------------- scoring
def test_normalize_weights_sums_to_100():
    w = scoring.normalize_weights({"skills": 50, "education": 50, "experience": 0,
                                   "projects": 0, "certifications": 0})
    assert round(sum(w.values())) == 100
    assert w["skills"] == w["education"] == 50.0


def test_normalize_weights_all_zero_falls_back_to_defaults():
    w = scoring.normalize_weights({k: 0 for k in config.CRITERIA})
    assert w == config.DEFAULT_WEIGHTS


def test_compute_shortlists_above_threshold():
    analysis = {"scores": {k: 80 for k in config.CRITERIA}}
    v = scoring.compute(analysis, scoring.normalize_weights(None),
                        scoring.normalize_cutoffs({k: 0 for k in config.CRITERIA}), 60)
    assert v["status"] == "SHORTLISTED"
    assert v["ats_score"] == 80.0


def test_compute_cutoff_forces_review():
    scores = {k: 80 for k in config.CRITERIA}
    scores["skills"] = 30  # below the cutoff, total still clears the threshold
    v = scoring.compute({"scores": scores}, scoring.normalize_weights(None),
                        scoring.normalize_cutoffs({"skills": 40}), 50)
    assert v["status"] == "REVIEW"
    assert v["failed_cutoffs"] == ["skills"]


def test_compute_below_threshold_not_shortlisted():
    v = scoring.compute({"scores": {k: 10 for k in config.CRITERIA}},
                        scoring.normalize_weights(None),
                        scoring.normalize_cutoffs({k: 0 for k in config.CRITERIA}), 60)
    assert v["status"] == "NOT_SHORTLISTED"


def test_build_candidate_fills_na():
    row = scoring.build_candidate("CID-X", "a.pdf", {}, scoring.normalize_weights(None),
                                  scoring.normalize_cutoffs(None), 60)
    assert row["candidate_name"] == "NA"
    assert row["email_id"] == "NA"
    assert row["ats_score"] == 0.0


def test_blank_candidate_is_manual_and_shortlisted():
    row = scoring.blank_candidate("CID-Y")
    assert row["manually_added"] is True
    assert row["status"] == "SHORTLISTED"


# ---------------------------------------------------------------- extractors
def test_extract_txt_roundtrip():
    out = extractors.extract_upload("resume.txt", ("Jane Doe\n" + "Python developer. " * 10).encode())
    assert len(out) == 1 and not out[0].error
    assert "Jane Doe" in out[0].text


def test_extract_zip_flattens_and_names():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("batch/jane.txt", "Jane Doe\n" + "Python developer. " * 10)
        zf.writestr("__MACOSX/junk.txt", "ignore me")
        zf.writestr("notes.exe", "not a resume")
    out = extractors.extract_upload("upload.zip", buf.getvalue())
    assert len(out) == 1
    assert out[0].file_name == "upload.zip/batch/jane.txt"
    assert not out[0].error


def test_extract_zip_entry_cap():
    old = config.MAX_ZIP_ENTRIES
    config.MAX_ZIP_ENTRIES = 2
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            for i in range(5):
                zf.writestr(f"r{i}.txt", "Some resume text long enough to parse. " * 5)
        out = extractors.extract_upload("many.zip", buf.getvalue())
        errors = [r for r in out if r.error]
        assert any("stopped after" in r.error for r in errors)
        assert len([r for r in out if not r.error]) <= 2
    finally:
        config.MAX_ZIP_ENTRIES = old


def test_extract_zip_member_size_cap():
    old = config.MAX_FILE_MB
    config.MAX_FILE_MB = 0.0001  # ~105 bytes
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("big.txt", "x" * 10_000)
        out = extractors.extract_upload("bomb.zip", buf.getvalue())
        assert any("per-file limit" in r.error for r in out)
    finally:
        config.MAX_FILE_MB = old


# ------------------------------------------------------------ interview links
def test_link_token_roundtrip_and_tamper():
    os.environ["INTERVIEW_LINK_SECRET"] = "test-secret"
    try:
        token = interview_link.make_token("SES-1", "CID-1")
        parsed = interview_link.parse_token(token)
        assert parsed and parsed["shortlist_id"] == "SES-1"
        assert parsed["candidate_id"] == "CID-1" and not parsed["expired"]
        # flip a signature character -> refused
        assert interview_link.parse_token(token[:-2] + ("AA" if token[-2:] != "AA" else "BB")) is None
    finally:
        del os.environ["INTERVIEW_LINK_SECRET"]


def test_link_token_expiry():
    os.environ["INTERVIEW_LINK_SECRET"] = "test-secret"
    try:
        token = interview_link.make_token("SES-1", "CID-1", days=-1)
        parsed = interview_link.parse_token(token)
        assert parsed and parsed["expired"] is True
    finally:
        del os.environ["INTERVIEW_LINK_SECRET"]


# ------------------------------------------------------------- link injection
def test_inject_link_placeholder():
    body, how = ai_agent.inject_link("Hi\n\n[INTERVIEW_LINK]\n\nBest regards,\nSam", "http://x/i/t")
    assert how == "placeholder" and "http://x/i/t" in body and "[INTERVIEW_LINK]" not in body


def test_inject_link_before_signoff_when_placeholder_missing():
    body, how = ai_agent.inject_link("Hi there.\n\nBest regards,\nSam", "http://x/i/t")
    assert how == "before-signoff"
    assert body.index("http://x/i/t") < body.index("Best regards")


def test_inject_link_appended_as_last_resort():
    body, how = ai_agent.inject_link("Hi there. No signoff here", "http://x/i/t")
    assert how == "appended" and body.endswith(("left off.",))


def test_inject_link_strips_placeholder_when_no_link():
    body, how = ai_agent.inject_link("Hi\n\n[INTERVIEW_LINK]\n\nBye", "")
    assert how == "none" and "[INTERVIEW_LINK]" not in body


def test_display_name_softens_block_capitals():
    assert ai_agent.display_name("PRIYA SUNDARAM") == "Priya Sundaram"
    assert ai_agent.display_name("NA") == "Candidate"


# ------------------------------------------------------------- fallback parse
def test_fallback_parse_extracts_contact():
    text = "Jane Doe\njane.doe@example.com\n+91 98765 43210\n5 years experience\nB.Tech in CS"
    data = ai_agent.fallback_parse(text, "jane_doe_resume.pdf")
    assert data["email_id"] == "jane.doe@example.com"
    assert data["experience_years"] == 5.0
    assert data["candidate_name"] == "Jane Doe"


# --------------------------------------------------------------- excel export
def test_excel_safe_neutralises_formulas():
    assert excel_export._safe("=HYPERLINK(\"http://evil\")").startswith("'")
    assert excel_export._safe("+SUM(A1)").startswith("'")
    assert excel_export._safe("@cmd").startswith("'")
    assert excel_export._safe("plain text") == "plain text"
    assert excel_export._safe(42) == 42


def test_excel_workbook_has_no_live_formula():
    from openpyxl import load_workbook
    record = {"candidates": [{"candidate_id": "CID-1", "candidate_name": "=HYPERLINK(\"x\")",
                              "status": "SHORTLISTED"}],
              "stats": {}, "weights": {}, "cutoffs": {}}
    data = excel_export.build_workbook(record)
    ws = load_workbook(io.BytesIO(data))["Shortlist"]
    cell = ws.cell(row=2, column=2)
    assert cell.data_type != "f", "candidate-controlled text must never become a formula"
    assert str(cell.value).lstrip("'").startswith("=HYPERLINK")


if __name__ == "__main__":
    failed = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"  ok    {name}")
            except AssertionError as exc:
                failed += 1
                print(f"  FAIL  {name}: {exc}")
            except Exception as exc:  # noqa: BLE001
                failed += 1
                print(f"  ERROR {name}: {type(exc).__name__}: {exc}")
    print(f"\n{failed} failure(s)" if failed else "\nall tests passed")
    sys.exit(1 if failed else 0)
