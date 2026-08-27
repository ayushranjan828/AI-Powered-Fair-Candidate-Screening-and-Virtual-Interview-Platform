"""Excel (.xlsx) export of a shortlist, with a second sheet documenting the run."""
from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (key, header, width) - the first seven are the columns required by the spec.
COLUMNS = [
    ("candidate_id", "Candidate ID", 18),
    ("candidate_name", "Candidate Name", 26),
    ("phone_number", "Phone Number", 18),
    ("email_id", "Email ID", 30),
    ("skills", "Skills", 55),
    ("certification", "Certification", 34),
    ("experience", "Experience", 40),
    ("highest_education", "Highest Education", 28),
    ("projects", "Projects", 45),
    ("ats_score", "ATS Score (%)", 14),
    ("score_education", "Education Score", 15),
    ("score_skills", "Skills Score", 13),
    ("score_experience", "Experience Score", 16),
    ("score_projects", "Projects Score", 14),
    ("score_certifications", "Certification Score", 18),
    ("status", "Status", 18),
    ("recommendation", "AI Recommendation", 18),
    ("justification", "AI Justification", 60),
    ("source_file", "Source Resume", 32),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SHORTLIST_FILL = PatternFill("solid", fgColor="E7F6EC")
REVIEW_FILL = PatternFill("solid", fgColor="FFF6E0")


def build_workbook(record: dict) -> bytes:
    candidates = record.get("candidates", [])
    wb = Workbook()

    ws = wb.active
    ws.title = "Shortlist"
    ws.append([c[1] for c in COLUMNS])
    for idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"

    for cand in candidates:
        ws.append([cand.get(key, "NA") for key, _, _ in COLUMNS])
        row = ws.max_row
        status = str(cand.get("status", ""))
        if status == "SHORTLISTED":
            fill = SHORTLIST_FILL
        elif status == "REVIEW":
            fill = REVIEW_FILL
        else:
            fill = None
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(ws.max_row, 1)}"

    # ---- run metadata sheet
    meta = wb.create_sheet("Screening Details")
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 95
    stats = record.get("stats", {})
    weights = record.get("weights", {})
    cutoffs = record.get("cutoffs", {})
    rows = [
        ("Job Title", record.get("job_title", "NA")),
        ("Session ID", record.get("session_id", "NA")),
        ("History ID", record.get("history_id", "-")),
        ("Generated / Accepted At", record.get("accepted_at") or record.get("created_at", "NA")),
        ("Accepted By", record.get("accepted_by", "-")),
        ("ATS Threshold (%)", record.get("threshold", "NA")),
        ("Criteria Weights", json.dumps(weights)),
        ("Criteria Cutoffs", json.dumps(cutoffs)),
        ("Resumes Received", stats.get("total", 0)),
        ("Successfully Parsed", stats.get("parsed", 0)),
        ("Failed To Parse", stats.get("failed", 0)),
        ("Shortlisted", stats.get("shortlisted", 0)),
        ("Rows In This Sheet", len(candidates)),
        ("Reviewer Notes", record.get("notes", "")),
        ("Job Description", (record.get("jd_text") or "")[:30000]),
    ]
    for label, value in rows:
        meta.append([label, value])
    for cell in meta["A"]:
        cell.font = Font(bold=True)
    for cell in meta["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
