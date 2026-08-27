"""Excel (.xlsx) export of one interview: report, parameter breakdown, transcript.

Three sheets so a reviewer can start at the summary and drill all the way down to
the exact words that produced each score.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config

HEADER_FILL = PatternFill("solid", fgColor="1F3B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GOOD_FILL = PatternFill("solid", fgColor="E7F6EC")
MID_FILL = PatternFill("solid", fgColor="FFF6E0")
BAD_FILL = PatternFill("solid", fgColor="FDECEB")

TRANSCRIPT_COLUMNS = [
    ("question_id", "Q#", 10),
    ("category_label", "Category", 20),
    ("question_source", "Planned / Follow-up", 18),
    ("difficulty", "Difficulty", 11),
    ("question", "Question asked", 58),
    ("answer", "Candidate answer", 70),
    ("words", "Words", 8),
    ("answer_seconds", "Seconds", 9),
    ("answer_type", "Answer type", 14),
    ("scores", "Scores given", 40),
    ("strengths", "Strengths noted", 40),
    ("concerns", "Concerns noted", 40),
    ("evidence", "Evidence quoted", 45),
    ("intent", "What it tested", 40),
]


def _fill_for(score) -> PatternFill | None:
    try:
        value = float(score)
    except (TypeError, ValueError):
        return None
    if value >= 70:
        return GOOD_FILL
    if value >= 50:
        return MID_FILL
    return BAD_FILL


def _style_header(ws, width_source) -> None:
    for idx, (_, _, width) in enumerate(width_source, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def _join(values) -> str:
    if isinstance(values, (list, tuple)):
        return "\n".join(f"- {v}" for v in values) if values else ""
    return str(values or "")


SUMMARY_COLUMNS = [
    ("candidate_name", "Candidate", 26),
    ("email_id", "Email", 30),
    ("current_role", "Current role", 34),
    ("attendance_label", "Attended?", 16),
    ("answered", "Questions answered", 19),
    ("overall_score", "Interview score (%)", 19),
    ("verdict", "AI verdict", 16),
    ("confidence", "Report confidence", 17),
    ("decision_label", "Human decision", 18),
    ("reviewer", "Reviewed by", 20),
    ("reviewed_at", "Reviewed at", 22),
    ("override_score", "Overridden score", 17),
    ("ats_score", "Resume ATS (%)", 16),
    ("stage_label", "Stage", 16),
    ("link_state", "Interview link", 20),
    ("interview_id", "Interview ID", 20),
]

ATTENDANCE_LABELS = {
    "COMPLETED": "Yes - completed",
    "PARTIAL": "Yes - did not finish",
    "NOT_STARTED": "No - never started",
    "NO_INTERVIEW": "No - not invited",
}

DECISION_LABELS = {
    "PROCEED": "Proceed",
    "HOLD": "Hold",
    "REJECT": "Do not proceed",
    "": "Not decided",
}


def build_summary_workbook(board: dict) -> bytes:
    """One row per shortlisted candidate: who attended, how they scored, and what
    a human decided.

    This is the report a recruiter actually circulates - the per-interview
    workbook from build_workbook() is the drill-down for one person.
    """
    rows = board.get("rows", [])
    wb = Workbook()

    ws = wb.active
    ws.title = "Candidates"
    ws.append([c[1] for c in SUMMARY_COLUMNS])
    _style_header(ws, SUMMARY_COLUMNS)
    ws.freeze_panes = "B2"

    for row in rows:
        iv = row.get("interview") or {}
        review = iv.get("human_review") or {}
        invite = row.get("invite") or {}
        link_state = ("withdrawn" if invite.get("revoked")
                      else "sent" if invite.get("sent_at")
                      else "issued" if invite else "none")
        values = {
            "candidate_name": row.get("candidate_name", "NA"),
            "email_id": row.get("email_id", "NA"),
            "current_role": row.get("current_role", "NA"),
            "attendance_label": ATTENDANCE_LABELS.get(row.get("attendance"), "NA"),
            "answered": iv.get("answered", 0),
            "overall_score": iv.get("overall_score") if iv.get("overall_score") is not None else "",
            "verdict": (iv.get("verdict") or "").replace("_", " ").title(),
            "confidence": iv.get("confidence", ""),
            "decision_label": DECISION_LABELS.get(row.get("decision", ""), "Not decided"),
            "reviewer": review.get("reviewer", ""),
            "reviewed_at": review.get("reviewed_at", ""),
            "override_score": review.get("override_score")
                              if review.get("override_score") is not None else "",
            "ats_score": row.get("ats_score", ""),
            "stage_label": (row.get("stage") or "").replace("_", " ").title(),
            "link_state": link_state,
            "interview_id": iv.get("interview_id", ""),
        }
        ws.append([values.get(key, "") for key, _, _ in SUMMARY_COLUMNS])
        line = ws.max_row
        for col in range(1, len(SUMMARY_COLUMNS) + 1):
            ws.cell(row=line, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        # Colour the attendance and score cells so the split is readable at a glance.
        if not row.get("attended"):
            ws.cell(row=line, column=4).fill = BAD_FILL
        elif row.get("attendance") == "PARTIAL":
            ws.cell(row=line, column=4).fill = MID_FILL
        else:
            ws.cell(row=line, column=4).fill = GOOD_FILL
        fill = _fill_for(iv.get("overall_score"))
        if fill:
            ws.cell(row=line, column=6).fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{max(ws.max_row, 1)}"

    # ---- run metadata --------------------------------------------------------
    stats = board.get("stats", {})
    attendance = stats.get("attendance", {})
    decisions = stats.get("decisions", {})
    meta = wb.create_sheet("Summary")
    meta.column_dimensions["A"].width = 34
    meta.column_dimensions["B"].width = 60
    for label, value in [
        ("Role", board.get("job_title", "NA")),
        ("Shortlist", board.get("history_id", "NA")),
        ("Shortlist accepted", board.get("accepted_at", "NA")),
        ("", ""),
        ("Shortlisted candidates", stats.get("total", 0)),
        ("Attended the interview", stats.get("attended", 0)),
        ("  of which completed", attendance.get("COMPLETED", 0)),
        ("  of which unfinished", attendance.get("PARTIAL", 0)),
        ("Did NOT attend", stats.get("not_attended", 0)),
        ("  invited, never started", attendance.get("NOT_STARTED", 0)),
        ("  never invited", attendance.get("NO_INTERVIEW", 0)),
        ("", ""),
        ("Decision: Proceed", decisions.get("PROCEED", 0)),
        ("Decision: Hold", decisions.get("HOLD", 0)),
        ("Decision: Do not proceed", decisions.get("REJECT", 0)),
        ("Decision: not recorded yet", decisions.get("NONE", 0)),
        ("", ""),
        ("Average interview score (%)",
         stats.get("average_score") if stats.get("average_score") is not None else "NA"),
        ("", ""),
        ("Note", "Resume ATS is the screening-stage score. It forms no part of the "
                 "interview score and the interviewer is never told it."),
    ]:
        meta.append([label, value])
    for cell in meta["A"]:
        cell.font = Font(bold=True)
    for cell in meta["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_workbook(interview: dict) -> bytes:
    report = interview.get("report") or {}
    candidate = interview.get("candidate") or {}
    parameters = report.get("parameters") or {}
    weights = report.get("parameter_weights") or {}
    notes = report.get("parameter_notes") or {}
    coverage = report.get("coverage") or {}

    wb = Workbook()

    # ---- 1. Report -----------------------------------------------------------
    ws = wb.active
    ws.title = "Interview Report"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100

    rows = [
        ("Candidate", candidate.get("candidate_name", "NA")),
        ("Email", candidate.get("email_id", "NA")),
        ("Current role", candidate.get("current_role", "NA")),
        ("Interviewed for", interview.get("job_title", "NA")),
        ("Interviewer", f"{interview.get('interviewer', {}).get('name', 'NA')} "
                        f"({interview.get('interviewer', {}).get('role', 'NA')})"),
        ("Interview ID", interview.get("interview_id", "NA")),
        ("Started", interview.get("started_at") or interview.get("created_at", "NA")),
        ("Completed", interview.get("completed_at", "NA")),
        ("", ""),
        ("OVERALL INTERVIEW SCORE", f"{report.get('overall_score', 'NA')} %"),
        ("Verdict", report.get("verdict", "NOT_ASSESSED")),
        ("Report confidence", report.get("confidence", "NA")),
        ("Why that confidence", _join(report.get("confidence_reasons"))),
        ("", ""),
        ("Summary", report.get("summary", "")),
        ("Recommended next step", report.get("recommended_next_step", "")),
        ("", ""),
        ("Strengths shown in the interview", _join(report.get("strengths"))),
        ("Gaps shown in the interview", _join(report.get("gaps"))),
        ("Standout moments", _join(report.get("standout_moments"))),
        ("Not covered by this interview", _join(report.get("not_covered"))),
        ("Risk flags", _join(report.get("risk_flags")) or "None"),
        ("", ""),
        ("Questions asked", coverage.get("asked", 0)),
        ("Questions answered", coverage.get("answered", 0)),
        ("Follow-up questions asked", coverage.get("followups", 0)),
        ("Answers graded", coverage.get("graded", 0)),
        ("Answers not graded", coverage.get("ungraded", 0)),
        ("Total words spoken", coverage.get("total_words", 0)),
        ("Total answering time (min)", round(float(coverage.get("total_seconds") or 0) / 60, 1)),
        ("", ""),
        ("Question plan source", (interview.get("plan") or {}).get("source", "NA")),
        ("Closing review source", report.get("review_source", "NA")),
        ("Notes recorded during the run", interview.get("plan_error", "") or "None"),
        ("", ""),
        ("Resume ATS score (screening stage - NOT part of this score)",
         (report.get("screening_reference") or {}).get("ats_score", "NA")),
        ("Job description", (interview.get("jd_text") or "")[:30000]),
    ]
    for label, value in rows:
        ws.append([label, value])
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    for cell in ws["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Make the headline number impossible to miss.
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        if row[0].value == "OVERALL INTERVIEW SCORE":
            row[0].font = Font(bold=True, size=13)
            row[1].font = Font(bold=True, size=13)
            fill = _fill_for(report.get("overall_score"))
            if fill:
                row[0].fill = fill
                row[1].fill = fill

    # ---- 2. Parameters -------------------------------------------------------
    param_cols = [
        ("parameter", "Parameter", 24),
        ("weight", "Weight (%)", 12),
        ("score", "Score (%)", 11),
        ("turn_score", "From per-answer grades", 22),
        ("holistic_score", "From closing review", 20),
        ("answers", "Answers behind it", 17),
        ("basis", "How it was derived", 42),
        ("note", "Interviewer note", 70),
        ("meaning", "What this parameter means", 50),
    ]
    ps = wb.create_sheet("Parameters")
    ps.append([c[1] for c in param_cols])
    _style_header(ps, param_cols)
    for key, meaning in config.PARAMETERS.items():
        entry = parameters.get(key) or {}
        ps.append([
            key.replace("_", " ").title(),
            weights.get(key, ""),
            entry.get("score") if entry.get("score") is not None else "not evidenced",
            entry.get("turn_score") if entry.get("turn_score") is not None else "-",
            entry.get("holistic_score") if entry.get("holistic_score") is not None else "-",
            entry.get("answers", 0),
            entry.get("basis", ""),
            notes.get(key, ""),
            meaning,
        ])
        row = ps.max_row
        fill = _fill_for(entry.get("score"))
        for col in range(1, len(param_cols) + 1):
            cell = ps.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and col == 3:
                cell.fill = fill
    ps.freeze_panes = "B2"

    # ---- 3. Transcript -------------------------------------------------------
    ts = wb.create_sheet("Transcript")
    ts.append([c[1] for c in TRANSCRIPT_COLUMNS])
    _style_header(ts, TRANSCRIPT_COLUMNS)

    for turn in interview.get("turns", []):
        assessment = turn.get("assessment") or {}
        metrics = turn.get("metrics") or {}
        scores = assessment.get("scores") or {}
        values = {
            **turn,
            "words": metrics.get("words", 0),
            "answer_type": assessment.get("answer_type", ""),
            "scores": "\n".join(f"{k.replace('_', ' ')}: {v}" for k, v in scores.items()),
            "strengths": _join(assessment.get("strengths")),
            "concerns": _join(assessment.get("concerns")),
            "evidence": assessment.get("evidence", ""),
            "question_source": "Follow-up" if turn.get("question_source") == "followup"
                               else "Planned",
        }
        ts.append([values.get(key, "") for key, _, _ in TRANSCRIPT_COLUMNS])
        row = ts.max_row
        for col in range(1, len(TRANSCRIPT_COLUMNS) + 1):
            ts.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        if turn.get("question_source") == "followup":
            ts.cell(row=row, column=3).fill = MID_FILL

    ts.freeze_panes = "B2"
    ts.auto_filter.ref = f"A1:{get_column_letter(len(TRANSCRIPT_COLUMNS))}{max(ts.max_row, 1)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
